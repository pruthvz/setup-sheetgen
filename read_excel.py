"""Read Book1.xlsx and PUNCH.xlsx and output contents as readable text."""
import openpyxl
from pathlib import Path

BASE = Path(__file__).parent


def format_cell(value):
    """Format cell value for display."""
    if value is None:
        return ""
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return f"{value:.4f}".rstrip("0").rstrip(".")
    return str(value)


def read_sheet(ws, max_rows=None):
    """Read worksheet and return formatted rows."""
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if max_rows and row_idx > max_rows:
            break
        cells = [format_cell(c) for c in row]
        while cells and cells[-1] == "":
            cells.pop()
        if cells:
            rows.append(cells)
    return rows


def output_sheet(name, rows, col_width=15):
    """Print sheet data in aligned columns."""
    if not rows:
        print("  (empty)")
        return
    ncols = max(len(r) for r in rows)
    widths = [col_width] * ncols
    for r in rows:
        for i, c in enumerate(r):
            if i < ncols:
                widths[i] = max(widths[i], min(len(str(c)) + 2, 30))
    for r in rows:
        line = "  ".join(str(c).ljust(widths[i]) for i, c in enumerate(r))
        print(line)


def main():
    print("=" * 80)
    print("Book1.xlsx")
    print("=" * 80)

    wb1 = openpyxl.load_workbook(BASE / "Book1.xlsx", read_only=True, data_only=True)
    for sheet_name in wb1.sheetnames:
        ws = wb1[sheet_name]
        print(f"\n--- Sheet: {sheet_name} ---\n")
        rows = read_sheet(ws, max_rows=250)
        output_sheet(sheet_name, rows)
        if len(rows) >= 250:
            print(f"\n  ... (truncated at 250 rows)")
    wb1.close()

    print("\n")
    print("=" * 80)
    print("PUNCH.xlsx")
    print("=" * 80)

    wb2 = openpyxl.load_workbook(BASE / "PUNCH.xlsx", read_only=True, data_only=True)
    for sheet_name in wb2.sheetnames:
        ws = wb2[sheet_name]
        print(f"\n--- Sheet: {sheet_name} ---\n")
        rows = read_sheet(ws, max_rows=None)
        output_sheet(sheet_name, rows)
    wb2.close()


if __name__ == "__main__":
    main()
